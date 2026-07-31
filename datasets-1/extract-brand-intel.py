"""Extracts the Altherma competitive-intelligence sheet into a typed TS module.

Reads datasets-1/"Daikin Competitor Intel (1).xlsx" (sheet Altherma), rows 16-84,
and writes src/data/brand-intel.ts.

Run from the repo root:  python datasets-1/extract-brand-intel.py
"""

import openpyxl

PATH = "datasets-1/Daikin Competitor Intel (1).xlsx"
SHEET = "Altherma"
OUT = "src/data/brand-intel.ts"

BRAND_COLS = [
    ("daikin", "E"),
    ("mitsubishi", "F"),
    ("panasonic", "G"),
    ("vaillant", "H"),
    ("viessmann", "I"),
    ("bosch", "J"),
]

CAT_SLUG = {
    "Executive": "executive",
    "Product Team": "product",
    "Sales": "sales",
    "Technology": "technology",
    "Marketing": "marketing",
    "R&D": "rnd",
    "Regulatory & Compliance": "regulatory",
    "Customer / Voice of Customer": "customer",
    "Pricing & Profitability": "pricing",
    "Financial Performance": "financial",
    "Competitive Positioning": "positioning",
    "Macro / Market Outlook": "macro",
    "Aftermarket & Services": "aftermarket",
}

REPLACEMENTS = {
    "–": "-", "—": "-", "’": "'", "‘": "'",
    "“": '"', "”": '"', "²": "2", "°": "deg",
    "→": "->", " ": " ",
}

HEADER = '''/* AUTO-GENERATED from datasets-1/"Daikin Competitor Intel (1).xlsx" (sheet Altherma).
 * Regenerate with datasets-1/extract-brand-intel.py -- do not edit by hand.
 *
 * IMPORTANT PROVENANCE NOTE
 * Every per-brand value in this file is SYNTHETIC: plausible, internally
 * consistent placeholders authored for demo/POC use. Nothing here is researched
 * or sourced, and none of it may be quoted externally or drive a real decision.
 * The `ingestion` and `visualization` fields are genuine recommendations, and
 * `suggestedSource` records where real data would have to come from.
 */

export const BRAND_INTEL_SYNTHETIC = true;
export const BRAND_INTEL_SCOPE =
  "EU/EMEA residential air-to-water heat pumps, FY2025 baseline";

export type BrandKey =
  | "daikin"
  | "mitsubishi"
  | "panasonic"
  | "vaillant"
  | "viessmann"
  | "bosch";

export interface BrandMeta {
  key: BrandKey;
  name: string;
  product: string;
  isDaikin: boolean;
  /** Validated categorical hue. Fixed order, never cycled or reassigned by rank. */
  color: string;
}

export const BRANDS: BrandMeta[] = [
  { key: "daikin", name: "Daikin", product: "Altherma", isDaikin: true, color: "#0097e0" },
  { key: "mitsubishi", name: "Mitsubishi Electric", product: "Ecodan", isDaikin: false, color: "#d1366b" },
  { key: "panasonic", name: "Panasonic", product: "Aquarea", isDaikin: false, color: "#78933c" },
  { key: "vaillant", name: "Vaillant", product: "aroTHERM", isDaikin: false, color: "#9c5ce0" },
  { key: "viessmann", name: "Viessmann", product: "Vitocal", isDaikin: false, color: "#00a878" },
  { key: "bosch", name: "Bosch/Buderus", product: "Compress", isDaikin: false, color: "#cc7a00" },
];

export const BRAND_BY_KEY: Record<BrandKey, BrandMeta> = Object.fromEntries(
  BRANDS.map((b) => [b.key, b]),
) as Record<BrandKey, BrandMeta>;

export type CategoryKey =
  | "executive"
  | "positioning"
  | "product"
  | "technology"
  | "rnd"
  | "sales"
  | "pricing"
  | "marketing"
  | "customer"
  | "aftermarket"
  | "regulatory"
  | "financial"
  | "macro";

export interface CategoryMeta {
  key: CategoryKey;
  label: string;
  /** Which internal team this category primarily serves. */
  audience: string;
  blurb: string;
}

export const CATEGORIES: CategoryMeta[] = [
  { key: "executive", label: "Executive", audience: "Leadership", blurb: "Market size, share, strategy and the risks that reach the board." },
  { key: "positioning", label: "Competitive Positioning", audience: "Leadership / Strategy", blurb: "Where each rival beats or trails Daikin, and how threatening they are." },
  { key: "product", label: "Product Team", audience: "Product", blurb: "Specs, certifications, lifecycle stage, roadmap and feature gaps." },
  { key: "technology", label: "Technology", audience: "Engineering", blurb: "Refrigerant roadmap, compressors, controls, cold-climate behaviour and IP." },
  { key: "rnd", label: "R&D", audience: "Engineering", blurb: "Spend, patent throughput, innovation pipeline and time-to-market." },
  { key: "sales", label: "Sales", audience: "Sales", blurb: "Revenue, price movement, channel mix, win rates and cycle length." },
  { key: "pricing", label: "Pricing & Profitability", audience: "Sales / Finance", blurb: "List positioning, discounting behaviour, margin and TCO claims." },
  { key: "marketing", label: "Marketing", audience: "Marketing", blurb: "Awareness, spend, digital footprint, events and sustainability messaging." },
  { key: "customer", label: "Voice of Customer", audience: "Marketing / Service", blurb: "Installer and end-user satisfaction, reliability, service and network." },
  { key: "aftermarket", label: "Aftermarket & Services", audience: "Service", blurb: "Service attach, remote monitoring, parts, warranty and training scale." },
  { key: "regulatory", label: "Regulatory & Compliance", audience: "Compliance", blurb: "F-Gas, efficiency mandates, subsidies, certification and risk exposure." },
  { key: "financial", label: "Financial Performance", audience: "Finance", blurb: "Segment profitability, capex, valuation and financial health." },
  { key: "macro", label: "Macro / Market Outlook", audience: "Strategy", blurb: "Demand drivers, construction cycle, policy tailwinds and forecasts." },
];

export const CATEGORY_BY_KEY: Record<CategoryKey, CategoryMeta> = Object.fromEntries(
  CATEGORIES.map((c) => [c.key, c]),
) as Record<CategoryKey, CategoryMeta>;

export interface IntelMetric {
  /** Source row in the spreadsheet -- keeps every value traceable to the sheet. */
  sourceRow: number;
  category: CategoryKey;
  metric: string;
  whatToCapture: string;
  visualization: string;
  ingestion: string;
  suggestedSource: string;
  values: Record<BrandKey, string>;
}

export const INTEL_METRICS: IntelMetric[] = [
'''

FOOTER = '''];

export function metricsForCategory(category: CategoryKey): IntelMetric[] {
  return INTEL_METRICS.filter((m) => m.category === category);
}

/** Ingestion mechanism family, for the "how this arrives" view. */
export function ingestionKind(ingestion: string): "api" | "manual" | "internal" | "mixed" {
  const v = ingestion.toLowerCase();
  const api = v.startsWith("api");
  const manual = v.startsWith("manual");
  if (v.startsWith("manual/api") || v.startsWith("api/manual")) return "mixed";
  if (v.startsWith("internal")) return "internal";
  if (api) return "api";
  if (manual) return "manual";
  return "internal";
}
'''


def ts(value):
    if value is None:
        return '""'
    text = str(value)
    for bad, good in REPLACEMENTS.items():
        text = text.replace(bad, good)
    text = text.replace("\\", "\\\\").replace('"', '\\"')
    text = " ".join(text.split())
    return '"' + text + '"'


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET]

    out = [HEADER]
    category = ""
    count = 0

    for row in range(16, 85):
        label = ws[f"A{row}"].value
        if label:
            category = label
        slug = CAT_SLUG[category]

        out.append("  {")
        out.append(f"    sourceRow: {row},")
        out.append(f'    category: "{slug}",')
        out.append(f"    metric: {ts(ws[f'B{row}'].value)},")
        out.append(f"    whatToCapture: {ts(ws[f'C{row}'].value)},")
        out.append(f"    visualization: {ts(ws[f'D{row}'].value)},")
        out.append(f"    ingestion: {ts(ws[f'K{row}'].value)},")
        out.append(f"    suggestedSource: {ts(ws[f'L{row}'].value)},")
        out.append("    values: {")
        for key, col in BRAND_COLS:
            out.append(f"      {key}: {ts(ws[f'{col}{row}'].value)},")
        out.append("    },")
        out.append("  },")
        count += 1

    out.append(FOOTER)

    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("\n".join(out))

    print(f"Wrote {count} metrics -> {OUT}")


if __name__ == "__main__":
    main()
